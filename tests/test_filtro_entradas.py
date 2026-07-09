"""Testes do filtro "Considerar apenas documentos de entrada no SPED".

Cobre a classificacao entrada/saida (IND_OPER e CFOP), o filtro nos motores
dos itens 1 (SPED x SEFAZ) e 2 (diff de SPEDs) e a indicacao do filtro nos
relatorios Excel.
"""

from __future__ import annotations

import os
import sys
import tempfile
from decimal import Decimal

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(RAIZ, "src"))

from openpyxl import load_workbook  # noqa: E402

from auditoria_fiscal.core.filtro_sped import (  # noqa: E402
    MSG_SEM_ENTRADAS, ROTULO_FILTRO_ENTRADAS, e_entrada,
    filtrar_documento_entradas, filtrar_entradas,
)
from auditoria_fiscal.core.modelos import (  # noqa: E402
    DocumentoFiscalConjunto, Empresa, ItemNota, NotaFiscal,
)
from auditoria_fiscal.ferramentas.comparador_sped_sefaz import comparar  # noqa: E402
from auditoria_fiscal.ferramentas.comparador_sped_sped import comparar_speds  # noqa: E402
from auditoria_fiscal.ferramentas.extracao_itens import (  # noqa: E402
    exportar_itens_excel, extrair_itens,
)
from auditoria_fiscal.ferramentas.relatorio_diff_excel import gerar_relatorio_diff  # noqa: E402
from auditoria_fiscal.ferramentas.relatorio_excel import gerar_relatorio  # noqa: E402


def chave(sufixo: str) -> str:
    ch = ("35260399888777000166550010000010011111111" + sufixo)[:44]
    return ch.ljust(44, "0")


def nota(ind_oper="", cfop="", numero="1", sufixo="1") -> NotaFiscal:
    n = NotaFiscal(chave=chave(sufixo), numero=numero, ind_oper=ind_oper,
                   valor_documento=Decimal("100"))
    if cfop:
        n.itens.append(ItemNota(num_item="1", cfop=cfop,
                                valor_item=Decimal("100")))
    return n


def test_classificacao() -> None:
    assert e_entrada(nota(ind_oper="0")) is True
    assert e_entrada(nota(ind_oper="1")) is False
    # Sem IND_OPER: decide pelo CFOP dos itens.
    assert e_entrada(nota(cfop="1102")) is True      # entrada estadual
    assert e_entrada(nota(cfop="2102")) is True      # entrada interestadual
    assert e_entrada(nota(cfop="3102")) is True      # entrada do exterior
    assert e_entrada(nota(cfop="5102")) is False     # venda estadual
    assert e_entrada(nota(cfop="6102")) is False     # venda interestadual
    assert e_entrada(nota(cfop="7102")) is False     # venda ao exterior
    # Sem nenhuma informacao: mantem a nota (compatibilidade).
    assert e_entrada(nota()) is True
    print("OK classificacao entrada/saida (IND_OPER + CFOP)")


def test_filtrar() -> None:
    notas = [nota("0", sufixo="1"), nota("1", sufixo="2"),
             nota(cfop="5102", sufixo="3"), nota(cfop="1102", sufixo="4")]
    entradas = filtrar_entradas(notas)
    assert [n.chave for n in entradas] == [notas[0].chave, notas[3].chave]

    doc = DocumentoFiscalConjunto(empresa=Empresa(nome="Teste"), notas=notas)
    doc_f = filtrar_documento_entradas(doc)
    assert len(doc_f.notas) == 2 and doc_f.empresa.nome == "Teste"
    assert len(doc.notas) == 4  # original intocado
    print("OK filtrar_entradas / filtrar_documento_entradas")


def test_diff_com_filtro() -> None:
    doc_a = DocumentoFiscalConjunto(notas=[
        nota("0", numero="1", sufixo="1"), nota("1", numero="2", sufixo="2")])
    doc_b = DocumentoFiscalConjunto(notas=[
        nota("0", numero="1", sufixo="1"), nota("1", numero="2", sufixo="2")])

    completo = comparar_speds(doc_a, doc_b)
    assert completo.total_a == 2 and completo.apenas_entradas is False

    so_entradas = comparar_speds(doc_a, doc_b, apenas_entradas=True)
    assert so_entradas.total_a == 1 and so_entradas.total_b == 1
    assert so_entradas.apenas_entradas is True

    # Relatorio indica o filtro no Resumo.
    with tempfile.TemporaryDirectory() as tmp:
        arq = os.path.join(tmp, "diff.xlsx")
        gerar_relatorio_diff(so_entradas, arq)
        ws = load_workbook(arq)["Resumo"]
        assert ws["A2"].value == ROTULO_FILTRO_ENTRADAS

        gerar_relatorio_diff(completo, arq)
        ws = load_workbook(arq)["Resumo"]
        assert ws["A2"].value != ROTULO_FILTRO_ENTRADAS
    print("OK diff SPED x SPED com apenas_entradas + relatorio")


def test_sefaz_com_filtro() -> None:
    doc = DocumentoFiscalConjunto(notas=[
        nota("0", sufixo="1"), nota("1", sufixo="2"),
        nota(cfop="5102", sufixo="3"),   # sem IND_OPER, saida pelo CFOP
    ])
    res = comparar(doc, [], apenas_entradas=True)
    assert res.total_sped == 1 and res.apenas_entradas is True

    res_todos = comparar(doc, [], apenas_entradas=False)
    assert res_todos.total_sped == 3 and res_todos.apenas_entradas is False

    with tempfile.TemporaryDirectory() as tmp:
        arq = os.path.join(tmp, "sefaz.xlsx")
        gerar_relatorio(res, arq, nome_empresa="Empresa Teste")
        ws = load_workbook(arq)["Resumo"]
        assert ws["A3"].value == ROTULO_FILTRO_ENTRADAS
        assert "somente entradas" in ws["A5"].value

        gerar_relatorio(res_todos, arq)
        ws = load_workbook(arq)["Resumo"]
        assert ws["A3"].value is None
        assert "todas as operacoes" in ws["A5"].value
    print("OK comparador SPED x SEFAZ com filtro + relatorio")


def test_extracao_com_filtro() -> None:
    notas = [nota("0", sufixo="1"), nota("1", sufixo="2")]
    notas[0].itens.append(ItemNota(num_item="1", cfop="1102",
                                   valor_item=Decimal("100")))
    notas[1].itens.append(ItemNota(num_item="1", cfop="5102",
                                   valor_item=Decimal("50")))

    linhas = extrair_itens(notas, somente_operacao="0")
    assert len(linhas) == 1 and linhas[0]["operacao"] == "Entrada"

    with tempfile.TemporaryDirectory() as tmp:
        arq = os.path.join(tmp, "itens.xlsx")
        exportar_itens_excel(linhas, arq, filtro_aplicado=ROTULO_FILTRO_ENTRADAS)
        ws = load_workbook(arq).active
        assert ws["A1"].value == ROTULO_FILTRO_ENTRADAS
        assert ws["A2"].value == "Chave de acesso"      # cabecalho desceu
        assert ws.auto_filter.ref.startswith("A2:")

        # Sem filtro: layout original preservado.
        exportar_itens_excel(linhas, arq)
        ws = load_workbook(arq).active
        assert ws["A1"].value == "Chave de acesso"
    print("OK extracao de itens com filtro + planilha")


def test_mensagens() -> None:
    assert "Nenhum documento de entrada" in MSG_SEM_ENTRADAS
    assert ROTULO_FILTRO_ENTRADAS.startswith("Filtro aplicado")
    print("OK textos padronizados")


if __name__ == "__main__":
    test_classificacao()
    test_filtrar()
    test_diff_com_filtro()
    test_sefaz_com_filtro()
    test_extracao_com_filtro()
    test_mensagens()
    print("\nTodos os testes do filtro de entradas passaram.")
