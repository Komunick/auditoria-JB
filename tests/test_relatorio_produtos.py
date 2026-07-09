"""Teste do relatorio Excel da auditoria de produtos (dados sinteticos)."""

from __future__ import annotations

import os
import sys
import tempfile
from decimal import Decimal

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(RAIZ, "src"))

from openpyxl import load_workbook  # noqa: E402

from auditoria_fiscal.core.cadastro_produtos import ProdutoCadastro  # noqa: E402
from auditoria_fiscal.ferramentas.auditoria_produtos import (  # noqa: E402
    CONF_ALTA, CONF_MEDIA, Inconsistencia, MSG_ST_COMO_TRIBUTADO,
    MSG_TRIBUTADO_COMO_ST, ResultadoAuditoria, TIPO_CEST_AUSENTE,
    TIPO_ST_COMO_TRIBUTADO, TIPO_TRIBUTADO_COMO_ST, TRIB_INTEGRAL, TRIB_ST,
)
from auditoria_fiscal.ferramentas.relatorio_produtos import (  # noqa: E402
    TITULOS_AUDITORIA, exportar_relatorio_excel,
)


def montar_resultados() -> list[ResultadoAuditoria]:
    p1 = ProdutoCadastro(indice=0, codigo="P001", descricao="BALA DE GOMA SORTIDA",
                         ncm="17041000", cest="1703100", cfops=["5405"],
                         cst="60", unidade="UN")
    r1 = ResultadoAuditoria(
        produto=p1, situacao="OK",
        tributacao_atual=TRIB_ST, tributacao_sugerida=TRIB_ST,
        fundamentacao="Anexo I RICMS/BA; Conv. ICMS 142/18 - CEST 17.031.00",
    )

    p2 = ProdutoCadastro(indice=1, codigo="P002", descricao="REFRIGERANTE COLA 2L",
                         ncm="22021000", cest="", cfops=["5102"], cst="00",
                         aliquota=Decimal("20.5"))
    r2 = ResultadoAuditoria(
        produto=p2, situacao="INCONSISTENTE",
        tributacao_atual=TRIB_INTEGRAL, tributacao_sugerida=TRIB_ST,
        confianca=CONF_ALTA,
        inconsistencias=[Inconsistencia(TIPO_ST_COMO_TRIBUTADO,
                                        MSG_ST_COMO_TRIBUTADO)],
        correcoes={"cst": "60", "cest": "0300200"},
        cfop_map={"5102": "5405"},
        fundamentacao="Anexo I RICMS/BA; Conv. ICMS 142/18 - CEST 03.002.00",
    )

    p3 = ProdutoCadastro(indice=2, codigo="P003",
                         descricao="SHAMPOO ANTICASPA 200ML",
                         ncm="33051000", cest="", cfops=["5405"], cst="60")
    r3 = ResultadoAuditoria(
        produto=p3, situacao="ALERTA",
        tributacao_atual=TRIB_ST, tributacao_sugerida=TRIB_ST,
        confianca=CONF_MEDIA,
        inconsistencias=[Inconsistencia(TIPO_CEST_AUSENTE,
                                        "Produto ST sem CEST informado.",
                                        nivel="alerta")],
        correcoes={"cest": "2001500"},
    )

    p4 = ProdutoCadastro(indice=3, codigo="P004", descricao="PARAFUSO SEXTAVADO M8",
                         ncm="73181500", cest="", cfops=["5405", "6404"],
                         cst="060")
    r4 = ResultadoAuditoria(
        produto=p4, situacao="INCONSISTENTE",
        tributacao_atual=TRIB_ST, tributacao_sugerida=TRIB_INTEGRAL,
        confianca=CONF_ALTA,
        inconsistencias=[Inconsistencia(TIPO_TRIBUTADO_COMO_ST,
                                        MSG_TRIBUTADO_COMO_ST)],
        correcoes={"cst": "000", "aliquota": "20,5"},
        cfop_map={"5405": "5102", "6404": "6102"},
        status_correcao="Corrigido",
    )
    return [r1, r2, r3, r4]


def cor_fundo(cel) -> str:
    cor = cel.fill.start_color
    return cor.rgb if isinstance(cor.rgb, str) else ""


def cor_fonte(cel) -> str:
    if cel.font is None or cel.font.color is None:
        return ""
    return cel.font.color.rgb if isinstance(cel.font.color.rgb, str) else ""


def achar_valor(ws, rotulo: str):
    """Procura o rotulo na coluna A e retorna o valor da coluna B."""
    for row in ws.iter_rows(min_col=1, max_col=2):
        if row[0].value == rotulo:
            return row[1].value
    return None


def main() -> int:
    resultados = montar_resultados()
    indicadores = {
        "total": 4, "corretos": 1, "inconsistentes": 2, "alertas": 1,
        "percentual_inconsistencias": 50.0, "sujeitos_st": 3,
        "st_incorretos": 2, "corrigidos": 1,
        "por_tipo": {TIPO_ST_COMO_TRIBUTADO: 1, TIPO_TRIBUTADO_COMO_ST: 1,
                     TIPO_CEST_AUSENTE: 1},
    }

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as fh:
        caminho = fh.name

    falhas: list[str] = []

    def checar(cond, msg):
        if not cond:
            falhas.append(msg)

    try:
        retorno = exportar_relatorio_excel(resultados, caminho, indicadores,
                                           contexto="cadastro_teste.xlsx")
        checar(retorno == caminho, f"retorno: {retorno}")
        checar(os.path.getsize(caminho) > 0, "arquivo vazio")

        wb = load_workbook(caminho)
        checar(wb.sheetnames == ["Resumo", "Auditoria"],
               f"abas: {wb.sheetnames}")

        # ---- Resumo ----
        ws = wb["Resumo"]
        checar(ws["A1"].value == "Relatorio de Auditoria Tributaria de Produtos",
               f"titulo: {ws['A1'].value}")
        checar(ws["A2"].value == "Arquivo: cadastro_teste.xlsx",
               f"contexto: {ws['A2'].value}")
        checar(str(ws["A3"].value or "").startswith("Gerado em: "),
               f"data/hora: {ws['A3'].value}")
        checar(achar_valor(ws, "Total analisados") == 4,
               f"Total analisados: {achar_valor(ws, 'Total analisados')}")
        checar(achar_valor(ws, "Inconsistentes") == 2,
               f"Inconsistentes: {achar_valor(ws, 'Inconsistentes')}")
        checar(achar_valor(ws, "% de inconsistencias") == 50.0,
               f"percentual: {achar_valor(ws, '% de inconsistencias')}")
        checar(achar_valor(ws, "Corrigidos") == 1,
               f"Corrigidos: {achar_valor(ws, 'Corrigidos')}")
        checar(achar_valor(ws, TIPO_ST_COMO_TRIBUTADO) == 1,
               "por_tipo: linha de ST como tributado ausente")
        checar(achar_valor(ws, TIPO_CEST_AUSENTE) == 1,
               "por_tipo: linha de CEST ausente ausente")

        # ---- Auditoria ----
        ws = wb["Auditoria"]
        cab = [c.value for c in ws[1]]
        checar(cab == TITULOS_AUDITORIA, f"cabecalho: {cab}")
        checar(ws.max_row == 1 + len(resultados),
               f"linhas: {ws.max_row} (esperado {1 + len(resultados)})")
        checar(ws.freeze_panes == "A2", f"freeze: {ws.freeze_panes}")
        checar(bool(ws.auto_filter.ref), "auto_filter ausente")

        col_sit = TITULOS_AUDITORIA.index("Situacao") + 1
        col_cfop = TITULOS_AUDITORIA.index("CFOP atual") + 1
        col_aliq = TITULOS_AUDITORIA.index("Aliquota") + 1
        col_tipo = TITULOS_AUDITORIA.index("Tipo da inconsistencia") + 1
        col_msg = TITULOS_AUDITORIA.index("Mensagens") + 1
        col_corr = TITULOS_AUDITORIA.index("Correcao sugerida") + 1
        col_status = TITULOS_AUDITORIA.index("Status") + 1

        # linha 2 = P001 (OK), 3 = P002 (INCONSISTENTE), 4 = P003 (ALERTA),
        # 5 = P004 (INCONSISTENTE corrigido)
        checar(ws.cell(row=2, column=1).value == "P001", "ordem das linhas")

        cel = ws.cell(row=3, column=col_sit)
        checar(cel.value == "INCONSISTENTE", f"situacao P002: {cel.value}")
        checar(cor_fundo(cel).endswith("FFC7CE"),
               f"fundo INCONSISTENTE: {cor_fundo(cel)}")
        checar(cor_fonte(cel).endswith("9C0006"),
               f"fonte INCONSISTENTE: {cor_fonte(cel)}")
        checar(cor_fundo(ws.cell(row=2, column=col_sit)).endswith("C6EFCE"),
               f"fundo OK: {cor_fundo(ws.cell(row=2, column=col_sit))}")
        checar(cor_fundo(ws.cell(row=4, column=col_sit)).endswith("FFEB9C"),
               f"fundo ALERTA: {cor_fundo(ws.cell(row=4, column=col_sit))}")

        checar(ws.cell(row=5, column=col_cfop).value == "5405, 6404",
               f"cfops P004: {ws.cell(row=5, column=col_cfop).value}")
        checar(ws.cell(row=3, column=col_aliq).value == "20,5",
               f"aliquota P002: {ws.cell(row=3, column=col_aliq).value}")
        checar(ws.cell(row=3, column=col_tipo).value == TIPO_ST_COMO_TRIBUTADO,
               f"tipo P002: {ws.cell(row=3, column=col_tipo).value}")
        checar(ws.cell(row=3, column=col_msg).value == MSG_ST_COMO_TRIBUTADO,
               f"mensagem P002: {ws.cell(row=3, column=col_msg).value}")

        corr2 = ws.cell(row=3, column=col_corr).value or ""
        checar("CST 00 -> 60" in corr2, f"correcao cst P002: {corr2}")
        checar("CFOP 5102 -> 5405" in corr2, f"correcao cfop P002: {corr2}")
        checar("CEST -> 0300200" in corr2, f"correcao cest P002: {corr2}")

        corr4 = ws.cell(row=5, column=col_corr).value or ""
        checar("CST 060 -> 000" in corr4, f"correcao cst P004: {corr4}")
        checar("Aliquota -> 20,5" in corr4, f"correcao aliquota P004: {corr4}")

        checar(ws.cell(row=5, column=col_status).value == "Corrigido",
               f"status P004: {ws.cell(row=5, column=col_status).value}")
        checar(ws.cell(row=2, column=col_status).value == "Nao corrigido",
               f"status P001: {ws.cell(row=2, column=col_status).value}")
        wb.close()

        # Sem indicadores: calcula via calcular_indicadores do motor.
        exportar_relatorio_excel(resultados, caminho)
        wb = load_workbook(caminho)
        checar(wb.sheetnames == ["Resumo", "Auditoria"],
               f"abas (indicadores None): {wb.sheetnames}")
        checar(achar_valor(wb["Resumo"], "Total analisados") == 4,
               "Total (indicadores None)")
        wb.close()
    finally:
        if os.path.exists(caminho):
            os.unlink(caminho)

    if falhas:
        print("FALHAS:")
        for f in falhas:
            print("  -", f)
        return 1
    print("OK - relatorio de auditoria de produtos passou.")
    print(f"  {len(resultados)} resultados | {len(TITULOS_AUDITORIA)} colunas | "
          "abas Resumo + Auditoria")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
