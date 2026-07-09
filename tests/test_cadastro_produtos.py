"""Teste do leitor/regravador da base de cadastro de produtos (autocontido)."""

from __future__ import annotations

import os
import shutil
import sys
import tempfile
from decimal import Decimal

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(RAIZ, "src"))

from openpyxl import Workbook, load_workbook  # noqa: E402

from auditoria_fiscal.core.cadastro_produtos import (  # noqa: E402
    gerar_nova_base, ler_base_produtos,
)

# Cabecalhos com acentos (montados via escapes p/ manter o fonte ASCII).
COL_CODIGO = "C\u00f3digo"                 # Codigo com o-agudo
COL_DESCRICAO = "Descri\u00e7\u00e3o"      # Descricao com c-cedilha e a-til
COL_ALIQ = "Al\u00edq. ICMS"               # Aliq com i-agudo


def montar_xlsx(caminho: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Produtos"
    ws.append(["Cadastro de Produtos"])
    ws.append([COL_CODIGO, COL_DESCRICAO, "NCM", "CEST", "CFOP", "CST",
               COL_ALIQ, "Grupo", "Un"])
    ws.append(["P001", "REFRIGERANTE COLA 2L", "2202.10.00", "03.002.00",
               5102, "00", 20.5, "BEBIDAS", "UN"])
    ws.append(["P002", "BALA DE GOMA 100G", "1704.90.10", "17.031.00",
               "5102/6102", "060", "18%", "DOCES", "UN"])
    ws.append([None] * 9)  # linha em branco no meio (invalida)
    ws.append(["P003", "PARAFUSO SEXTAVADO M8", "73181500", "",
               "5405", "60", "", "FERRAGENS", "UN"])
    ws.append(["P004", "SHAMPOO ERVAS 350ML", "3305.10.00", "20.015.00",
               "5405", "60", "", "HIGIENE", "UN"])
    ws.append(["TOTAL", "", "", "", "", "", "", "", ""])  # rodape
    wb.save(caminho)


CSV_LINHAS = [
    "C\u00f3digo;Descri\u00e7\u00e3o;NCM;CFOP;CST",
    "P1;REFRIGERANTE LATA;22021000;5102;00",
    "P2;PARAFUSO M8;73181500;5405;60",
]


def main() -> int:
    falhas = []

    def checar(cond, msg):
        if not cond:
            falhas.append(msg)

    pasta = tempfile.mkdtemp(prefix="teste_cadastro_")
    try:
        # ------------------------------------------------------------------
        # Cenario xlsx: cabecalho na linha 2, acentos, cfop duplo, rodape.
        # ------------------------------------------------------------------
        arq_xlsx = os.path.join(pasta, "cadastro.xlsx")
        montar_xlsx(arq_xlsx)
        base = ler_base_produtos(arq_xlsx)

        checar(base.layout.tipo == "xlsx", f"tipo: {base.layout.tipo}")
        checar(base.layout.nome_aba == "Produtos",
               f"aba: {base.layout.nome_aba}")
        checar(base.layout.linha_cabecalho == 1,
               f"linha_cabecalho: {base.layout.linha_cabecalho}")

        mapa = base.layout.mapa_colunas
        esperado = {"codigo": 0, "descricao": 1, "ncm": 2, "cest": 3,
                    "cfop": 4, "cst": 5, "aliquota": 6, "grupo": 7,
                    "unidade": 8}
        for campo, col in esperado.items():
            checar(mapa.get(campo) == col,
                   f"mapa[{campo}]: {mapa.get(campo)} (esperado {col})")
        checar(base.layout.colunas_cfop == [4],
               f"colunas_cfop: {base.layout.colunas_cfop}")

        produtos = base.produtos
        checar(len(produtos) == 4, f"produtos: {len(produtos)} (esperado 4)")
        indices = [p.indice for p in produtos]
        checar(indices == [0, 1, 3, 4],
               f"indices nao preservados: {indices} (esperado [0, 1, 3, 4])")

        p1, p2, p3, p4 = produtos
        checar(p1.codigo == "P001", f"p1 codigo: {p1.codigo}")
        checar(p1.ncm == "22021000", f"p1 ncm so digitos: {p1.ncm}")
        checar(p1.cest == "0300200", f"p1 cest so digitos: {p1.cest}")
        checar(p1.cfops == ["5102"], f"p1 cfops: {p1.cfops}")
        checar(p1.cst == "00", f"p1 cst: {p1.cst}")
        checar(p1.aliquota == Decimal("20.5"), f"p1 aliquota: {p1.aliquota}")
        checar(p1.grupo == "BEBIDAS", f"p1 grupo: {p1.grupo}")
        checar(p1.unidade == "UN", f"p1 unidade: {p1.unidade}")

        checar(p2.cfops == ["5102", "6102"], f"p2 cfops multiplos: {p2.cfops}")
        checar(p2.cst == "060", f"p2 cst preserva zeros: {p2.cst}")
        checar(p2.aliquota == Decimal("18"), f"p2 aliquota com %: {p2.aliquota}")

        checar(p3.ncm == "73181500", f"p3 ncm: {p3.ncm}")
        checar(p3.aliquota is None, f"p3 aliquota vazia: {p3.aliquota}")
        checar(p4.cest == "2001500", f"p4 cest: {p4.cest}")

        codigos = [p.codigo for p in produtos]
        checar("TOTAL" not in codigos, f"rodape TOTAL nao ignorado: {codigos}")
        checar(base.diagnostico["produtos_validos"] == 4,
               f"diagnostico produtos_validos: "
               f"{base.diagnostico['produtos_validos']}")

        # ------------------------------------------------------------------
        # gerar_nova_base (xlsx): altera cst + cfop_map SO do produto indice 1.
        # ------------------------------------------------------------------
        saida_xlsx = os.path.join(pasta, "cadastro_corrigido.xlsx")
        retorno = gerar_nova_base(base, saida_xlsx, {
            1: {"cst": "000", "cfop_map": {"5102": "5405", "6102": "6404"}},
        })
        checar(retorno == saida_xlsx, f"retorno gerar_nova_base: {retorno}")

        wb2 = load_workbook(saida_xlsx)
        ws2 = wb2["Produtos"]
        # Celulas alvo (linha ws 4 = produto indice 1).
        checar(ws2.cell(4, 5).value == "5405/6404",
               f"cfop_map com separador: {ws2.cell(4, 5).value}")
        checar(ws2.cell(4, 6).value == "000",
               f"cst alterado: {ws2.cell(4, 6).value}")
        # Demais celulas intactas.
        checar(ws2.cell(1, 1).value == "Cadastro de Produtos",
               f"titulo da linha 1 mudou: {ws2.cell(1, 1).value}")
        checar(ws2.cell(2, 1).value == COL_CODIGO,
               f"cabecalho mudou: {ws2.cell(2, 1).value}")
        checar(str(ws2.cell(3, 5).value) == "5102",
               f"cfop do P001 mudou: {ws2.cell(3, 5).value}")
        checar(ws2.cell(3, 6).value == "00",
               f"cst do P001 mudou: {ws2.cell(3, 6).value}")
        checar(ws2.cell(4, 1).value == "P002",
               f"codigo do P002 mudou: {ws2.cell(4, 1).value}")
        checar(ws2.cell(4, 3).value == "1704.90.10",
               f"ncm do P002 mudou: {ws2.cell(4, 3).value}")
        checar(ws2.cell(6, 5).value == "5405",
               f"cfop do P003 mudou: {ws2.cell(6, 5).value}")
        checar(ws2.cell(7, 6).value == "60",
               f"cst do P004 mudou: {ws2.cell(7, 6).value}")
        checar(ws2.cell(8, 1).value == "TOTAL",
               f"rodape mudou: {ws2.cell(8, 1).value}")
        wb2.close()

        # ------------------------------------------------------------------
        # Cenario csv (sep ";", latin-1): le, altera, regrava, confere texto.
        # ------------------------------------------------------------------
        arq_csv = os.path.join(pasta, "cadastro.csv")
        with open(arq_csv, "w", encoding="latin-1", newline="") as fh:
            fh.write("\n".join(CSV_LINHAS) + "\n")

        base_csv = ler_base_produtos(arq_csv)
        checar(base_csv.layout.tipo == "csv", f"csv tipo: {base_csv.layout.tipo}")
        checar(base_csv.layout.separador == ";",
               f"csv separador: {base_csv.layout.separador!r}")
        checar(base_csv.layout.encoding == "latin-1",
               f"csv encoding: {base_csv.layout.encoding}")
        checar(base_csv.layout.linha_cabecalho == 0,
               f"csv linha_cabecalho: {base_csv.layout.linha_cabecalho}")
        checar(len(base_csv.produtos) == 2,
               f"csv produtos: {len(base_csv.produtos)}")
        c1 = base_csv.produtos[0]
        checar(c1.codigo == "P1" and c1.ncm == "22021000"
               and c1.cfops == ["5102"] and c1.cst == "00",
               f"csv p1: {c1}")

        saida_csv = os.path.join(pasta, "cadastro_corrigido.csv")
        # "aliquota" sem coluna no arquivo: deve ser ignorada em silencio.
        gerar_nova_base(base_csv, saida_csv, {
            0: {"cst": "060", "cfop_map": {"5102": "5405"},
                "aliquota": "20,5"},
        })
        with open(saida_csv, "rb") as fh:
            texto_saida = fh.read().decode("latin-1")
        linhas_saida = texto_saida.splitlines()
        checar(len(linhas_saida) == 3, f"csv saida linhas: {len(linhas_saida)}")
        checar(linhas_saida[0] == CSV_LINHAS[0],
               f"csv cabecalho mudou: {linhas_saida[0]!r}")
        checar(linhas_saida[1] == "P1;REFRIGERANTE LATA;22021000;5405;060",
               f"csv linha alterada: {linhas_saida[1]!r}")
        checar(linhas_saida[2] == CSV_LINHAS[2],
               f"csv linha intacta mudou: {linhas_saida[2]!r}")
    finally:
        shutil.rmtree(pasta, ignore_errors=True)

    if falhas:
        print("FALHAS:")
        for f in falhas:
            print("  -", f)
        return 1
    print("OK - cadastro de produtos passou.")
    print(f"  xlsx: {len(produtos)} produtos | mapa {len(mapa)} colunas | "
          f"cfops p2 {p2.cfops}")
    print(f"  csv: {len(base_csv.produtos)} produtos | sep "
          f"{base_csv.layout.separador!r} | {base_csv.layout.encoding}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
