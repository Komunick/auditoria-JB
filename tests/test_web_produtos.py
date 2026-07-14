"""Versao web — Auditoria de Produtos de ponta a ponta.

Sobe o app FastAPI com TestClient e exercita: upload da planilha xlsx do
cadastro, auditoria como job contra bases legais sinteticas (colocadas em
dados_web/dados, que o servidor prefere a pasta dados/ do repo), previa
filtravel com indice estavel por produto, correcao por selecao e em lote
(alta confianca), historico em dados_web/historico_produtos.csv, relatorio
Excel e nova base corrigida preservando o layout original. Reusa os
construtores de test_auditoria_produtos.py e test_cadastro_produtos.py.
"""

from __future__ import annotations

import io
import os
import sys
import tempfile
import time

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(RAIZ, "src"))

_TMP = tempfile.mkdtemp(prefix="auditoria_web_prod_")
os.environ["AUDITORIA_WEB_DADOS"] = _TMP

from fastapi.testclient import TestClient  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

from auditoria_fiscal.ferramentas.correcao_produtos import (  # noqa: E402
    CABECALHO_HISTORICO,
)
from auditoria_fiscal.web.servidor import criar_app  # noqa: E402
from test_auditoria_produtos import montar_dados  # noqa: E402
from test_cadastro_produtos import montar_xlsx  # noqa: E402

_MEDIA_XLSX = ("application/vnd.openxmlformats-officedocument."
               "spreadsheetml.sheet")


def checar(cond, msg):
    if not cond:
        print(f"FALHOU - {msg}")
        raise SystemExit(1)


def esperar_job(cliente, job_id: str) -> dict:
    """Espera o job terminar e devolve o job inteiro (status + resultado)."""
    for _ in range(200):
        job = cliente.get(f"/api/jobs/{job_id}").json()
        if job["status"] in ("concluido", "erro"):
            return job
        time.sleep(0.1)
    raise AssertionError("job nao concluiu a tempo")


def resultados(cliente, sessao: str, filtro: str = "todos") -> dict:
    r = cliente.get(f"/api/produtos/resultados?sessao_id={sessao}"
                    f"&filtro={filtro}")
    checar(r.status_code == 200, f"resultados falhou: {r.status_code} {r.text}")
    return r.json()


def por_codigo(dados: dict) -> dict:
    return {item["codigo"]: item for item in dados["itens"]}


def main() -> int:
    # Bases legais SINTETICAS em dados_web/dados (mesmo conteudo do teste
    # do motor) para a auditoria ser deterministica.
    pasta_dados = os.path.join(_TMP, "dados")
    os.makedirs(pasta_dados, exist_ok=True)
    montar_dados(pasta_dados)

    cliente = TestClient(criar_app())

    # Sem login: API bloqueada
    checar(cliente.get("/api/produtos/resultados?sessao_id=x").status_code == 401,
           "API deveria exigir login")

    r = cliente.post("/api/bootstrap", json={
        "usuario": "weslley", "nome": "Weslley", "senha": "segredo1"})
    checar(r.status_code == 200, f"bootstrap falhou: {r.text}")

    # ------------------------------------------------------------------
    # Sessao + upload da planilha do cadastro (xlsx com titulo, linha em
    # branco e rodape TOTAL — indices de dados 0, 1, 3 e 4)
    sessao = cliente.post("/api/sessoes",
                          json={"ferramenta": "produtos"}).json()["sessao_id"]

    # Auditar antes do upload -> job termina em erro
    job = cliente.post("/api/produtos/auditar",
                       json={"sessao_id": sessao}).json()
    job = esperar_job(cliente, job["job_id"])
    checar(job["status"] == "erro" and "planilha" in job["erro"],
           f"auditar sem upload deveria dar erro: {job}")

    arq_xlsx = os.path.join(_TMP, "cadastro.xlsx")
    montar_xlsx(arq_xlsx)
    with open(arq_xlsx, "rb") as fh:
        conteudo = fh.read()
    r = cliente.post(f"/api/produtos/upload?sessao_id={sessao}",
                     files={"arquivo": ("cadastro.xlsx", conteudo,
                                        _MEDIA_XLSX)})
    checar(r.status_code == 200, f"upload falhou: {r.text}")

    # Resultados antes de auditar -> 422
    checar(cliente.get(
        f"/api/produtos/resultados?sessao_id={sessao}").status_code == 422,
        "resultados sem auditoria deveria dar 422")

    # ------------------------------------------------------------------
    # Auditoria (job)
    job = cliente.post("/api/produtos/auditar",
                       json={"sessao_id": sessao}).json()
    job = esperar_job(cliente, job["job_id"])
    checar(job["status"] == "concluido", f"auditoria falhou: {job['erro']}")
    resultado = job["resultado"]
    checar(resultado["total"] == 4, f"total auditado: {resultado['total']}")
    ind = resultado["indicadores"]
    checar(ind["inconsistentes"] >= 2,
           f"esperava >= 2 inconsistentes: {ind['inconsistentes']}")
    checar(ind["corrigidos"] == 0, f"corrigidos iniciais: {ind['corrigidos']}")
    checar(resultado["mapa_colunas"].get("ncm") == "NCM",
           f"mapa de colunas: {resultado['mapa_colunas']}")

    # Filtro invalido -> 422
    checar(cliente.get(
        f"/api/produtos/resultados?sessao_id={sessao}&filtro=x"
        ).status_code == 422, "filtro invalido deveria dar 422")

    # ------------------------------------------------------------------
    # Previa: indices estaveis, situacoes e correcoes sugeridas
    dados = resultados(cliente, sessao)
    checar(len(dados["itens"]) == 4, f"itens na previa: {len(dados['itens'])}")
    itens = por_codigo(dados)
    checar([itens[c]["indice"] for c in ("P001", "P002", "P003", "P004")]
           == [0, 1, 3, 4],
           f"indices estaveis: {[i['indice'] for i in dados['itens']]}")
    checar(dados["contexto"] == "cadastro.xlsx",
           f"contexto: {dados['contexto']}")

    p1 = itens["P001"]   # refrigerante ST vendido como tributado (alta)
    checar(p1["situacao"] == "INCONSISTENTE", f"P001 situacao: {p1['situacao']}")
    checar(p1["confianca"] == "alta", f"P001 confianca: {p1['confianca']}")
    checar(p1["tem_correcao"], "P001 deveria ter correcao sugerida")
    checar("CST 00 -> 60" in p1["correcao"] and "CFOP 5102 -> 5405"
           in p1["correcao"], f"P001 correcao: {p1['correcao']}")
    checar(p1["status"] == "Nao corrigido", f"P001 status: {p1['status']}")

    p3 = itens["P003"]   # parafuso tributado vendido como ST (alta)
    checar(p3["situacao"] == "INCONSISTENTE", f"P003 situacao: {p3['situacao']}")
    checar("CST 60 -> 00" in p3["correcao"] and "CFOP 5405 -> 5102"
           in p3["correcao"], f"P003 correcao: {p3['correcao']}")
    checar("Aliquota -> 20,5" in p3["correcao"],
           f"P003 aliquota sugerida: {p3['correcao']}")

    checar(itens["P004"]["situacao"] == "OK",
           f"P004 (shampoo ST correto): {itens['P004']['situacao']}")

    # Filtros do combo
    dados_inc = resultados(cliente, sessao, "inconsistentes")
    checar(all(i["situacao"] == "INCONSISTENTE" for i in dados_inc["itens"]),
           "filtro inconsistentes trouxe outras situacoes")
    dados_alta = resultados(cliente, sessao, "alta_confianca")
    codigos_alta = sorted(por_codigo(dados_alta))
    checar("P001" in codigos_alta and "P003" in codigos_alta,
           f"alta confianca deveria ter P001 e P003: {codigos_alta}")

    # ------------------------------------------------------------------
    # Corrigir selecionados (indice estavel do P001)
    r = cliente.post("/api/produtos/corrigir", json={
        "sessao_id": sessao, "indices": [p1["indice"]]})
    checar(r.status_code == 200, f"corrigir falhou: {r.text}")
    r = r.json()
    checar(r["corrigidos_agora"] == 1, f"corrigidos agora: {r}")
    checar(r["acumuladas"] == 1, f"acumuladas: {r}")
    checar(r["indicadores"]["corrigidos"] == 1,
           f"indicador corrigidos: {r['indicadores']}")

    dados = resultados(cliente, sessao)
    checar(por_codigo(dados)["P001"]["status"] == "Corrigido",
           "P001 deveria constar como Corrigido")

    # Corrigir o mesmo indice de novo -> 422 (nada pendente)
    checar(cliente.post("/api/produtos/corrigir", json={
        "sessao_id": sessao, "indices": [p1["indice"]]}).status_code == 422,
        "recorrigir o mesmo produto deveria dar 422")

    # Corrigir alta confianca (pega o P003, pulando o P001 ja corrigido)
    r = cliente.post("/api/produtos/corrigir", json={
        "sessao_id": sessao, "alta_confianca": True})
    checar(r.status_code == 200, f"corrigir alta falhou: {r.text}")
    r = r.json()
    checar(r["corrigidos_agora"] >= 1, f"alta confianca: {r}")
    dados = resultados(cliente, sessao)
    checar(por_codigo(dados)["P003"]["status"] == "Corrigido",
           "P003 deveria constar como Corrigido")
    acumuladas = r["acumuladas"]

    # Historico auditavel em dados_web (nunca no LOCALAPPDATA do processo)
    historico = os.path.join(_TMP, "historico_produtos.csv")
    checar(os.path.isfile(historico), "historico_produtos.csv nao gravado")
    with open(historico, encoding="utf-8-sig") as fh:
        linhas_hist = [ln for ln in fh.read().splitlines() if ln]
    checar(linhas_hist[0] == ";".join(CABECALHO_HISTORICO),
           f"cabecalho do historico: {linhas_hist[0]}")
    checar(any(";cst;" in ln for ln in linhas_hist[1:]),
           "linha de cst ausente no historico")

    # ------------------------------------------------------------------
    # Relatorio Excel (todas as linhas, nao so a previa)
    r = cliente.post(f"/api/produtos/relatorio?sessao_id={sessao}")
    checar(r.status_code == 200, f"relatorio falhou: {r.status_code}")
    checar(r.content.startswith(b"PK"), "relatorio nao e um xlsx (PK)")
    checar("auditoria_produtos.xlsx" in r.headers.get("content-disposition", ""),
           "filename do relatorio errado")

    # ------------------------------------------------------------------
    # Nova base corrigida (reabre o upload original e preserva o layout)
    r = cliente.post(f"/api/produtos/nova-base?sessao_id={sessao}")
    checar(r.status_code == 200, f"nova base falhou: {r.status_code} {r.text[:120]}")
    checar(r.content.startswith(b"PK"), "nova base nao e um xlsx (PK)")
    checar("cadastro_corrigida.xlsx" in r.headers.get("content-disposition", ""),
           f"filename da nova base: {r.headers.get('content-disposition')}")

    wb = load_workbook(io.BytesIO(r.content))
    ws = wb["Produtos"]
    checar(ws.cell(1, 1).value == "Cadastro de Produtos",
           "titulo original nao preservado")
    # P001 (linha 3): CST 00 -> 60 e CFOP 5102 -> 5405
    checar(ws.cell(3, 6).value == "60", f"cst P001: {ws.cell(3, 6).value}")
    checar(str(ws.cell(3, 5).value) == "5405", f"cfop P001: {ws.cell(3, 5).value}")
    # P003 (linha 6): CST 60 -> 00, CFOP 5405 -> 5102 e aliquota 20,5
    checar(ws.cell(6, 6).value == "00", f"cst P003: {ws.cell(6, 6).value}")
    checar(str(ws.cell(6, 5).value) == "5102", f"cfop P003: {ws.cell(6, 5).value}")
    checar(ws.cell(6, 7).value == "20,5", f"aliquota P003: {ws.cell(6, 7).value}")
    # P004 (linha 7) intacto
    checar(ws.cell(7, 6).value == "60", f"cst P004 mudou: {ws.cell(7, 6).value}")
    wb.close()

    # Nova base sem correcao acumulada -> 422 (sessao nova)
    sessao2 = cliente.post("/api/sessoes",
                           json={"ferramenta": "produtos"}).json()["sessao_id"]
    cliente.post(f"/api/produtos/upload?sessao_id={sessao2}",
                 files={"arquivo": ("cadastro.xlsx", conteudo, _MEDIA_XLSX)})
    job = cliente.post("/api/produtos/auditar",
                       json={"sessao_id": sessao2}).json()
    esperar_job(cliente, job["job_id"])
    checar(cliente.post(
        f"/api/produtos/nova-base?sessao_id={sessao2}").status_code == 422,
        "nova base sem correcoes deveria dar 422")

    print("OK - auditoria de produtos web (upload, job, previa filtravel, "
          f"correcoes com {acumuladas} produto(s) acumulado(s), historico, "
          "relatorio e nova base) passou.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
