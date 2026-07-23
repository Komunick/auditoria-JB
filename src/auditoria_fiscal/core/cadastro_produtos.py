"""Leitor e regravador da base de cadastro de produtos do cliente.

Le a planilha do cadastro de produtos (xlsx/xls/csv/txt) com deteccao
automatica de aba, linha de cabecalho e colunas - mesma tecnica tolerante de
`sefaz_relacao` (palavras-chave sem acento, mapeamento aproximado) -
PRESERVANDO o layout original em `df_bruto`/`LayoutBase` para permitir
regravar depois uma nova base corrigida:

  * xlsx: a regravacao abre o ARQUIVO ORIGINAL com openpyxl e edita apenas
    as celulas alteradas, preservando formatacao, formulas e tipos das
    demais celulas;
  * csv/txt: regrava o df bruto inteiro com o mesmo separador/encoding.
"""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field
from decimal import Decimal

import pandas as pd

from .modelos import so_digitos
from .utils import para_decimal


# ----------------------------------------------------------------------
# Modelos
# ----------------------------------------------------------------------
@dataclass
class LayoutBase:
    """Layout detectado do arquivo, guardado para a regravacao fiel."""

    tipo: str                    # "xlsx" | "csv"
    separador: str = ";"         # csv/txt
    encoding: str = "latin-1"    # csv/txt
    nome_aba: str = ""           # xlsx: aba usada
    linha_cabecalho: int = 0     # indice da linha de cabecalho no df bruto
    cabecalho: list[str] = field(default_factory=list)
    mapa_colunas: dict[str, int] = field(default_factory=dict)  # campo -> indice col
    colunas_cfop: list[int] = field(default_factory=list)       # TODAS as colunas com "cfop" no nome


@dataclass
class ProdutoCadastro:
    """Um produto lido da base do cliente (campos ja normalizados)."""

    indice: int                  # indice da linha de DADOS (0-based, apos o cabecalho)
    codigo: str = ""
    descricao: str = ""
    ncm: str = ""                # so digitos
    cest: str = ""               # so digitos
    cfops: list[str] = field(default_factory=list)   # todos os CFOPs achados na linha
    cst: str = ""                # texto original so digitos (ex. "060", "60", "102")
    aliquota: Decimal | None = None                  # None quando vazio/sem coluna
    grupo: str = ""
    unidade: str = ""


@dataclass
class BaseProdutos:
    """Base completa: layout + df bruto + produtos validos + diagnostico."""

    caminho: str
    layout: LayoutBase
    df_bruto: pd.DataFrame       # arquivo inteiro, header=None, dtype=str
    produtos: list[ProdutoCadastro]
    diagnostico: dict            # mapa de colunas legivel, totais, avisos


# ----------------------------------------------------------------------
# Auxiliares de texto
# ----------------------------------------------------------------------
_RE_FLOAT_INTEIRO = re.compile(r"-?\d+\.0")
_RE_CFOP = re.compile(r"\b[1-7]\d{3}\b")


def _sem_acento(texto: str) -> str:
    if not texto:
        return ""
    nfkd = unicodedata.normalize("NFKD", str(texto))
    return "".join(c for c in nfkd if not unicodedata.combining(c)).lower().strip()


def _texto(valor) -> str:
    """Converte celula do df em texto limpo ('' p/ vazio/NaN; tira '.0')."""
    if valor is None:
        return ""
    if isinstance(valor, float) and valor != valor:  # NaN
        return ""
    texto = str(valor).strip()
    if texto.lower() in ("nan", "none"):
        return ""
    # Numeros lidos do Excel como float inteiro ("5102.0" -> "5102").
    if _RE_FLOAT_INTEIRO.fullmatch(texto):
        texto = texto[:-2]
    return texto


# ----------------------------------------------------------------------
# Deteccao de aba / cabecalho / colunas
# ----------------------------------------------------------------------
# Palavras-chave que identificam cada campo no cabecalho (mais especifica
# primeiro; o mapeamento ordena por tamanho decrescente, como em sefaz_relacao).
_MAPA_COLUNAS = {
    "codigo": ["codigo do produto", "cod produto", "cod. produto", "cod item",
               "codigo interno", "codigo", "cod.", "cod", "referencia",
               "id produto", "sku"],
    "descricao": ["descricao do produto", "nome do produto", "descricao",
                  "produto", "desc"],
    "ncm": ["ncm", "classificacao fiscal", "class fiscal", "class. fiscal",
            "cod ncm"],
    "cest": ["cest"],
    "cfop": ["cfop"],
    "cst": ["cst/csosn", "csosn", "cst icms", "situacao tributaria",
            "sit trib", "sit. trib", "cst"],
    "aliquota": ["aliquota icms", "aliq icms", "% icms", "aliquota", "aliq",
                 "icms"],
    "grupo": ["grupo tributario", "grupo trib", "grupo de tributacao",
              "tributacao", "grupo"],
    "unidade": ["unidade", "und", "un.", "un", "um"],
}

# Palavras usadas para pontuar linhas candidatas a cabecalho (e abas).
_PALAVRAS_CABECALHO = {"codigo", "descricao", "produto", "ncm", "cest", "cfop",
                       "cst", "csosn", "aliq", "icms", "grupo", "unidade",
                       "referencia", "sku", "item", "tributacao", "desc"}


def _melhor_aba(xls: "pd.ExcelFile") -> str:
    """Escolhe a aba do cadastro de produtos em planilhas de varias abas."""
    nomes = xls.sheet_names
    if len(nomes) == 1:
        return nomes[0]
    for nome in nomes:
        if any(p in _sem_acento(nome) for p in ("produt", "cadastro", "item")):
            return nome
    melhor, melhor_pont = nomes[0], -1
    for nome in nomes:
        df = pd.read_excel(xls, sheet_name=nome, header=None, dtype=str, nrows=30)
        pont = 0
        for idx in range(len(df)):
            celulas = [_sem_acento(c) for c in df.iloc[idx].tolist() if c is not None]
            pont = max(pont, sum(1 for cel in celulas
                                 if any(p in cel for p in _PALAVRAS_CABECALHO)))
        if pont > melhor_pont:
            melhor_pont, melhor = pont, nome
    return melhor


def _carregar_csv(caminho: str) -> tuple[pd.DataFrame, str, str]:
    """Carrega csv/txt testando separadores e encodings; retorna (df, sep, enc)."""
    for encoding in ("utf-8-sig", "latin-1"):
        for sep in (";", ",", "\t", "|"):
            try:
                df = pd.read_csv(caminho, header=None, dtype=str, sep=sep,
                                 encoding=encoding, engine="python",
                                 keep_default_na=False, skip_blank_lines=False)
            except Exception:  # noqa: BLE001
                continue
            if df.shape[1] > 1:
                return df, sep, encoding
    df = pd.read_csv(caminho, header=None, dtype=str, sep=";",
                     encoding="latin-1", engine="python",
                     keep_default_na=False, skip_blank_lines=False)
    return df, ";", "latin-1"


def _detectar_linha_cabecalho(df: pd.DataFrame) -> int:
    """Acha a linha que parece ser o cabecalho (contem palavras-chave)."""
    melhor_idx, melhor_pont = 0, -1
    for idx in range(min(len(df), 30)):  # cabecalho costuma estar no topo
        celulas = [_sem_acento(c) for c in df.iloc[idx].tolist() if c is not None]
        pont = sum(1 for cel in celulas
                   if cel and any(p in cel for p in _PALAVRAS_CABECALHO))
        if pont > melhor_pont:
            melhor_pont, melhor_idx = pont, idx
    return melhor_idx if melhor_pont > 0 else 0


def _mapear_colunas(cabecalho: list[str]) -> tuple[dict[str, int], list[int]]:
    """Associa campos a indices de coluna; retorna (mapa, colunas_cfop)."""
    normalizado = [_sem_acento(c) for c in cabecalho]
    mapa: dict[str, int] = {}
    usados: set[int] = set()
    for campo, chaves in _MAPA_COLUNAS.items():
        # Ordena chaves da mais especifica (maior) para a mais curta.
        for alvo in sorted(chaves, key=len, reverse=True):
            for i, nome in enumerate(normalizado):
                if i in usados or not nome:
                    continue
                if nome == alvo or alvo in nome:
                    mapa[campo] = i
                    usados.add(i)
                    break
            if campo in mapa:
                break
    # TODAS as colunas com "cfop" no nome; a primeira e o mapa oficial.
    colunas_cfop = [i for i, nome in enumerate(normalizado) if "cfop" in nome]
    if colunas_cfop:
        mapa["cfop"] = colunas_cfop[0]
    else:
        mapa.pop("cfop", None)
    return mapa, colunas_cfop


# ----------------------------------------------------------------------
# Leitura
# ----------------------------------------------------------------------
def _extrair_cfops(linha: list, colunas_cfop: list[int]) -> list[str]:
    """Extrai todos os CFOPs (4 digitos, 1-7) das colunas CFOP, sem repetir."""
    cfops: list[str] = []
    for col in colunas_cfop:
        if col >= len(linha):
            continue
        for cfop in _RE_CFOP.findall(_texto(linha[col])):
            if cfop not in cfops:
                cfops.append(cfop)
    return cfops


def _eh_rodape(codigo: str, descricao: str, ncm: str, cest: str,
               cfops: list[str], cst: str) -> bool:
    """Linha de total/rodape: texto 'total...' sem nenhum campo fiscal."""
    texto = _sem_acento(codigo) or _sem_acento(descricao)
    if not texto.startswith(("total", "subtotal")):
        return False
    return not (ncm or cest or cfops or cst)


def ler_base_produtos(caminho: str) -> BaseProdutos:
    """Le a base de produtos do cliente. Retorna BaseProdutos completa.

    O diagnostico traz o mapeamento de colunas legivel, totais e avisos,
    util para conferir se a deteccao automatica acertou o layout.
    """
    lower = caminho.lower()
    if lower.endswith((".xlsx", ".xlsm", ".xls")):
        xls = pd.ExcelFile(caminho)
        aba = _melhor_aba(xls)
        df_bruto = pd.read_excel(xls, sheet_name=aba, header=None, dtype=str)
        layout = LayoutBase(tipo="xlsx", nome_aba=aba)
    else:
        df_bruto, sep, enc = _carregar_csv(caminho)
        layout = LayoutBase(tipo="csv", separador=sep, encoding=enc)

    linha_cab = _detectar_linha_cabecalho(df_bruto)
    return _montar_base(caminho, df_bruto, layout, linha_cab)


def ler_base_produtos_fdb(caminho: str, tabela: str) -> BaseProdutos:
    """Le uma TABELA de um banco Firebird (.FDB) como base de produtos.

    Usa o leitor compartilhado (core.fdb_reader) para trazer a tabela como
    texto, monta um df_bruto com o cabecalho na primeira linha (nomes reais
    das colunas do Firebird) e reaproveita EXATAMENTE o mesmo mapeamento por
    palavra-chave do fluxo de planilha. A nova base corrigida sai em CSV
    (nao se regrava no .FDB): o layout fica como csv/utf-8 para isso.
    """
    from . import fdb_reader

    # Toda leitura roda isolada num subprocesso (ver fdb_reader): um .fdb
    # corrompido nao derruba o servidor.
    #
    # DESEMPENHO: tabelas de ERP tem dezenas/centenas de colunas, mas a
    # auditoria so usa ~9. Primeiro pegamos SO os nomes das colunas (barato,
    # sem dados), mapeamos os campos, e entao lemos APENAS as colunas
    # necessarias — em tabela larga isso corta a leitura em varias vezes.
    nomes = fdb_reader.colunas_tabela(caminho, tabela)
    cabecalho = [_texto(c) for c in nomes]
    mapa, colunas_cfop = _mapear_colunas(cabecalho)
    indices = sorted(set(mapa.values()) | set(colunas_cfop))

    layout = LayoutBase(tipo="csv", separador=";", encoding="utf-8-sig",
                        nome_aba=tabela)
    if not indices:
        # Nenhuma coluna reconhecida: le tudo para o diagnostico ser honesto
        # (o usuario ve que o layout nao bateu).
        _c, linhas, truncado = fdb_reader.ler_tabela(caminho, tabela)
        df_bruto = pd.DataFrame([cabecalho] + linhas, dtype=str) if cabecalho \
            else pd.DataFrame()
    else:
        sel = [nomes[i] for i in indices]
        _c, linhas, truncado = fdb_reader.ler_tabela(caminho, tabela,
                                                     colunas=sel)
        # df_bruto = cabecalho (nomes reais das colunas lidas) + dados. O
        # _montar_base remapeia sobre esse subconjunto — os mesmos campos, so
        # que nas novas posicoes.
        df_bruto = pd.DataFrame([[_texto(n) for n in sel]] + linhas, dtype=str)
    # No FDB o cabecalho e SEMPRE a primeira linha (nomes das colunas), sem a
    # deteccao heuristica que planilhas baguncadas exigem.
    base = _montar_base(caminho, df_bruto, layout, linha_cab=0)
    if truncado:
        base.diagnostico["avisos"].insert(
            0, f"Tabela grande: apenas as primeiras {len(linhas)} linhas foram "
               "lidas do .FDB (as demais foram ignoradas).")
    return base


def _montar_base(caminho: str, df_bruto: "pd.DataFrame", layout: LayoutBase,
                 linha_cab: int) -> BaseProdutos:
    """Nucleo comum: do df_bruto + layout, mapeia colunas e extrai produtos."""
    if df_bruto.empty:
        cabecalho: list[str] = []
    else:
        cabecalho = [_texto(c) for c in df_bruto.iloc[linha_cab].tolist()]
    mapa, colunas_cfop = _mapear_colunas(cabecalho)

    layout.linha_cabecalho = linha_cab
    layout.cabecalho = cabecalho
    layout.mapa_colunas = mapa
    layout.colunas_cfop = colunas_cfop

    def pega(linha: list, campo: str) -> str:
        idx = mapa.get(campo)
        if idx is None or idx >= len(linha):
            return ""
        return _texto(linha[idx])

    # Materializa as linhas de dados de UMA vez (df.values.tolist()) em vez de
    # df.iloc[i] a cada iteracao — iloc por linha e O(n) escondido e domina o
    # tempo em base grande (100k+ linhas de um .FDB).
    dados = df_bruto.iloc[linha_cab + 1:].values.tolist() if not df_bruto.empty \
        else []
    produtos: list[ProdutoCadastro] = []
    for indice, linha in enumerate(dados):
        codigo = pega(linha, "codigo")
        descricao = pega(linha, "descricao")
        # Linha de dados valida = tem codigo OU descricao nao vazios.
        if not codigo and not descricao:
            continue
        ncm = so_digitos(pega(linha, "ncm"))
        cest = so_digitos(pega(linha, "cest"))
        cst = so_digitos(pega(linha, "cst"))
        cfops = _extrair_cfops(linha, colunas_cfop)
        if _eh_rodape(codigo, descricao, ncm, cest, cfops, cst):
            continue
        texto_aliq = pega(linha, "aliquota").replace("%", "").strip()
        aliquota = para_decimal(texto_aliq) if texto_aliq else None
        produtos.append(ProdutoCadastro(
            indice=indice,
            codigo=codigo,
            descricao=descricao,
            ncm=ncm,
            cest=cest,
            cfops=cfops,
            cst=cst,
            aliquota=aliquota,
            grupo=pega(linha, "grupo"),
            unidade=pega(linha, "unidade"),
        ))

    avisos: list[str] = []
    if "codigo" not in mapa and "descricao" not in mapa:
        avisos.append("Colunas de codigo e descricao nao identificadas - "
                      "verifique o cabecalho do arquivo.")
    if "ncm" not in mapa:
        avisos.append("Coluna NCM nao identificada.")
    if "cst" not in mapa:
        avisos.append("Coluna CST/CSOSN nao identificada.")
    if not colunas_cfop:
        avisos.append("Nenhuma coluna CFOP identificada.")
    if not produtos:
        avisos.append("Nenhuma linha de dados valida encontrada.")

    def nome_col(i: int) -> str:
        return cabecalho[i] if i < len(cabecalho) and cabecalho[i] else f"col{i}"

    diagnostico = {
        "tipo": layout.tipo,
        "aba": layout.nome_aba,
        "separador": layout.separador,
        "encoding": layout.encoding,
        "linha_cabecalho": linha_cab,
        "cabecalho_detectado": cabecalho,
        "mapa_colunas": {campo: nome_col(i) for campo, i in mapa.items()},
        "colunas_cfop": [nome_col(i) for i in colunas_cfop],
        "total_linhas_dados": len(dados),
        "produtos_validos": len(produtos),
        "avisos": avisos,
    }
    return BaseProdutos(caminho=caminho, layout=layout, df_bruto=df_bruto,
                        produtos=produtos, diagnostico=diagnostico)


# ----------------------------------------------------------------------
# Regravacao
# ----------------------------------------------------------------------
_CAMPOS_SIMPLES = ("ncm", "cest", "cst", "aliquota")


def _aplicar_cfop_map(texto: str, cfop_map: dict[str, str]) -> str:
    """Troca cada CFOP de 4 digitos da celula pelo mapeado (preserva separadores)."""
    return _RE_CFOP.sub(lambda m: str(cfop_map.get(m.group(0), m.group(0))), texto)


def _gerar_xlsx(base: BaseProdutos, caminho_saida: str,
                alteracoes: dict[int, dict[str, object]]) -> None:
    """Edita o arquivo ORIGINAL via openpyxl (preserva formatacao) e salva."""
    from openpyxl import load_workbook

    layout = base.layout
    wb = load_workbook(base.caminho,
                       keep_vba=base.caminho.lower().endswith(".xlsm"))
    ws = wb[layout.nome_aba] if layout.nome_aba in wb.sheetnames else wb.active
    for indice, campos in alteracoes.items():
        linha_ws = layout.linha_cabecalho + 1 + int(indice) + 1  # ws e 1-based
        for campo in _CAMPOS_SIMPLES:
            if campo not in campos:
                continue
            col = layout.mapa_colunas.get(campo)
            if col is None:
                continue  # coluna inexistente no layout: ignora
            ws.cell(row=linha_ws, column=col + 1, value=str(campos[campo]))
        cfop_map = campos.get("cfop_map") or {}
        if not cfop_map:
            continue
        for col in layout.colunas_cfop:
            cel = ws.cell(row=linha_ws, column=col + 1)
            texto = "" if cel.value is None else str(cel.value)
            if not texto:
                continue
            novo = _aplicar_cfop_map(texto, cfop_map)
            if novo != texto:
                cel.value = novo
    wb.save(caminho_saida)


def _gerar_csv(base: BaseProdutos, caminho_saida: str,
               alteracoes: dict[int, dict[str, object]]) -> None:
    """Regrava o df bruto (com alteracoes) no mesmo separador/encoding."""
    layout = base.layout
    df = base.df_bruto.copy()
    primeira_linha = layout.linha_cabecalho + 1
    for indice, campos in alteracoes.items():
        lin = primeira_linha + int(indice)
        if lin >= len(df):
            continue
        for campo in _CAMPOS_SIMPLES:
            if campo not in campos:
                continue
            col = layout.mapa_colunas.get(campo)
            if col is None or col >= df.shape[1]:
                continue  # coluna inexistente no layout: ignora
            df.iat[lin, col] = str(campos[campo])
        cfop_map = campos.get("cfop_map") or {}
        if not cfop_map:
            continue
        for col in layout.colunas_cfop:
            if col >= df.shape[1]:
                continue
            texto = _texto(df.iat[lin, col])
            if not texto:
                continue
            novo = _aplicar_cfop_map(texto, cfop_map)
            if novo != texto:
                df.iat[lin, col] = novo
    df = df.fillna("")
    df.to_csv(caminho_saida, sep=layout.separador, encoding=layout.encoding,
              header=False, index=False)


def gerar_nova_base(base: BaseProdutos, caminho_saida: str,
                    alteracoes: dict[int, dict[str, object]]) -> str:
    """Gera a nova base corrigida preservando o layout original.

    `alteracoes[indice]` = {"ncm": str, "cest": str, "cst": str,
    "aliquota": str, "cfop_map": dict[str, str]} - todas as chaves opcionais.
    Campos cuja coluna nao existe no layout sao ignorados silenciosamente.
    Retorna `caminho_saida`.
    """
    if base.layout.tipo == "xlsx":
        _gerar_xlsx(base, caminho_saida, alteracoes)
    else:
        _gerar_csv(base, caminho_saida, alteracoes)
    return caminho_saida
