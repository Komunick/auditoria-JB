"""Item 5 - Relatorio Excel da auditoria tributaria do cadastro de produtos.

Exporta os resultados do motor (ferramentas.auditoria_produtos) em duas abas:
"Resumo" (indicadores gerais e contagem por tipo de inconsistencia) e
"Auditoria" (uma linha por produto, com a coluna Situacao destacada em cores).
"""

from __future__ import annotations

from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .auditoria_produtos import ResultadoAuditoria, calcular_indicadores

_AZUL = "1F4E78"

# situacao -> (cor de fundo, cor da fonte) — padrao dos estilos "Ruim/Neutro/Bom".
_CORES_SITUACAO: dict[str, tuple[str, str]] = {
    "INCONSISTENTE": ("FFC7CE", "9C0006"),
    "ALERTA": ("FFEB9C", "9C6500"),
    "OK": ("C6EFCE", "006100"),
}

TITULOS_AUDITORIA: list[str] = [
    "Codigo", "Produto", "NCM", "CEST", "CFOP atual", "CST/CSOSN", "Aliquota",
    "Tributacao atual", "Tributacao sugerida", "Confianca", "Situacao",
    "Tipo da inconsistencia", "Mensagens", "Correcao sugerida",
    "Fundamentacao legal", "Status",
]

_LARGURAS_AUDITORIA = [12, 40, 12, 11, 16, 11, 10, 24, 24, 11, 16, 34, 50, 36, 44, 14]

# (rotulo exibido, chave em indicadores)
_ITENS_RESUMO: list[tuple[str, str]] = [
    ("Total analisados", "total"),
    ("Corretos", "corretos"),
    ("Inconsistentes", "inconsistentes"),
    ("Alertas", "alertas"),
    ("% de inconsistencias", "percentual_inconsistencias"),
    ("Sujeitos a ST", "sujeitos_st"),
    ("ST incorretos", "st_incorretos"),
    ("Corrigidos", "corrigidos"),
]


def _cabecalho(ws, linha: int, titulos: list[str]) -> None:
    """Escreve uma linha de cabecalho no estilo padrao do projeto."""
    fonte = Font(bold=True, color="FFFFFF")
    fundo = PatternFill("solid", fgColor=_AZUL)
    for col, titulo in enumerate(titulos, start=1):
        cel = ws.cell(row=linha, column=col, value=titulo)
        cel.font = fonte
        cel.fill = fundo
        cel.alignment = Alignment(horizontal="center", vertical="center")


def _ajustar_larguras(ws, larguras: list[int]) -> None:
    for i, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = largura


def _texto_aliquota(valor) -> str:
    """Aliquota Decimal -> texto BR ("20,5"); None/vazio -> ""."""
    if valor is None or valor == "":
        return ""
    return str(valor).replace(".", ",")


def _texto_correcao(resultado: ResultadoAuditoria) -> str:
    """Monta o texto legivel da correcao sugerida.

    Ex.: "CST 000 -> 060; CFOP 5102 -> 5405; CEST -> 0300200; Aliquota -> 20,5".
    """
    partes: list[str] = []
    correcoes = resultado.correcoes or {}
    produto = resultado.produto
    if "cst" in correcoes:
        atual = produto.cst
        if atual:
            partes.append(f"CST {atual} -> {correcoes['cst']}")
        else:
            partes.append(f"CST -> {correcoes['cst']}")
    for de, para in (resultado.cfop_map or {}).items():
        partes.append(f"CFOP {de} -> {para}")
    if "cest" in correcoes:
        atual = produto.cest
        if atual:
            partes.append(f"CEST {atual} -> {correcoes['cest']}")
        else:
            partes.append(f"CEST -> {correcoes['cest']}")
    if "aliquota" in correcoes:
        atual = _texto_aliquota(produto.aliquota)
        if atual:
            partes.append(f"Aliquota {atual} -> {correcoes['aliquota']}")
        else:
            partes.append(f"Aliquota -> {correcoes['aliquota']}")
    return "; ".join(partes)


def _montar_resumo(ws, indicadores: dict, contexto: str) -> None:
    ws["A1"] = "Relatorio de Auditoria Tributaria de Produtos"
    ws["A1"].font = Font(bold=True, size=14, color=_AZUL)

    linha = 2
    if contexto:
        ws.cell(row=linha, column=1, value=f"Arquivo: {contexto}")
        linha += 1
    ws.cell(row=linha, column=1,
            value="Gerado em: " + datetime.now().strftime("%d/%m/%Y %H:%M"))
    linha += 2

    _cabecalho(ws, linha, ["Indicador", "Valor"])
    for i, (rotulo, chave) in enumerate(_ITENS_RESUMO, start=1):
        r = linha + i
        ws.cell(row=r, column=1, value=rotulo).font = Font(bold=True)
        cel = ws.cell(row=r, column=2, value=indicadores.get(chave, 0))
        if chave == "percentual_inconsistencias":
            cel.number_format = '0.0"%"'
    linha += len(_ITENS_RESUMO) + 2

    _cabecalho(ws, linha, ["Por tipo de inconsistencia", "Quantidade"])
    por_tipo = indicadores.get("por_tipo") or {}
    ordenados = sorted(por_tipo.items(), key=lambda kv: (-kv[1], kv[0]))
    for i, (tipo, quantidade) in enumerate(ordenados, start=1):
        ws.cell(row=linha + i, column=1, value=tipo)
        ws.cell(row=linha + i, column=2, value=quantidade)

    _ajustar_larguras(ws, [46, 14])


def _montar_auditoria(ws, resultados: list[ResultadoAuditoria]) -> None:
    _cabecalho(ws, 1, TITULOS_AUDITORIA)
    ws.freeze_panes = "A2"
    col_situacao = TITULOS_AUDITORIA.index("Situacao") + 1

    for resultado in resultados:
        produto = resultado.produto
        ws.append([
            produto.codigo,
            produto.descricao,
            produto.ncm,
            produto.cest,
            ", ".join(produto.cfops),
            produto.cst,
            _texto_aliquota(produto.aliquota),
            resultado.tributacao_atual,
            resultado.tributacao_sugerida,
            resultado.confianca,
            resultado.situacao,
            resultado.tipos,
            "; ".join(i.mensagem for i in resultado.inconsistencias),
            _texto_correcao(resultado),
            resultado.fundamentacao,
            resultado.status_correcao,
        ])
        cores = _CORES_SITUACAO.get(resultado.situacao)
        if cores:
            cel = ws.cell(row=ws.max_row, column=col_situacao)
            cel.fill = PatternFill("solid", fgColor=cores[0])
            cel.font = Font(color=cores[1])

    _ajustar_larguras(ws, _LARGURAS_AUDITORIA)
    if ws.max_row > 1:
        ultima = get_column_letter(len(TITULOS_AUDITORIA))
        ws.auto_filter.ref = f"A1:{ultima}{ws.max_row}"


def exportar_relatorio_excel(resultados: list[ResultadoAuditoria], caminho: str,
                             indicadores: dict | None = None,
                             contexto: str = "") -> str:
    """Exporta o relatorio da auditoria de produtos em Excel. Retorna o caminho."""
    if indicadores is None:
        indicadores = calcular_indicadores(resultados)

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"
    _montar_resumo(ws, indicadores, contexto)

    ws = wb.create_sheet("Auditoria")
    _montar_auditoria(ws, resultados)

    wb.save(caminho)
    return caminho
