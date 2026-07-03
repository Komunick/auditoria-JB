"""Exportacao do resultado da comparacao SPED x SEFAZ para Excel (.xlsx)."""

from __future__ import annotations

from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .comparador_sped_sefaz import ResultadoComparacao


_AZUL = "1F4E78"
_CINZA = "D9D9D9"
_VERMELHO = "C00000"
_AMARELO = "FFF2CC"


def _cabecalho(ws, titulos: list[str]) -> None:
    fonte = Font(bold=True, color="FFFFFF")
    fundo = PatternFill("solid", fgColor=_AZUL)
    for col, titulo in enumerate(titulos, start=1):
        cel = ws.cell(row=1, column=col, value=titulo)
        cel.font = fonte
        cel.fill = fundo
        cel.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def _ajustar_larguras(ws, larguras: list[int]) -> None:
    for i, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = largura


def _num(valor: Decimal) -> float:
    return float(valor or 0)


def _data(dt) -> str:
    return dt.strftime("%d/%m/%Y") if dt else ""


def gerar_relatorio(resultado: ResultadoComparacao, caminho: str,
                    nome_empresa: str = "") -> str:
    """Gera o arquivo Excel do relatorio de conferencia. Retorna o caminho."""
    wb = Workbook()

    # ---------------- Aba Resumo ----------------
    ws = wb.active
    ws.title = "Resumo"
    ws["A1"] = "Conferencia SPED x SEFAZ"
    ws["A1"].font = Font(bold=True, size=14, color=_AZUL)
    if nome_empresa:
        ws["A2"] = nome_empresa
        ws["A2"].font = Font(bold=True, size=11)

    resumo = resultado.resumo()
    linhas = [
        ("Notas na relacao da SEFAZ", resumo["notas_na_sefaz"]),
        ("Notas escrituradas no SPED (entradas)", resumo["notas_no_sped"]),
        ("Conciliadas (presentes em ambos)", resumo["conciliadas"]),
        ("FALTANTES no SPED (na SEFAZ, nao escrituradas)", resumo["faltantes_no_sped"]),
        ("Canceladas/denegadas porem escrituradas", resumo["canceladas_escrituradas"]),
        ("Divergencias de valor", resumo["divergencias_valor"]),
        ("Apenas no SPED (sem correspondencia na SEFAZ)", resumo["apenas_no_sped"]),
    ]
    inicio = 4
    for i, (rotulo, valor) in enumerate(linhas):
        r = inicio + i
        ws.cell(row=r, column=1, value=rotulo).font = Font(bold=True)
        cel = ws.cell(row=r, column=2, value=valor)
        # Destaca as pendencias principais.
        if rotulo.startswith("FALTANTES") and valor:
            cel.font = Font(bold=True, color=_VERMELHO)
    _ajustar_larguras(ws, [48, 14])

    # ---------------- Aba Faltantes no SPED ----------------
    ws = wb.create_sheet("Faltantes no SPED")
    _cabecalho(ws, ["Chave de acesso", "Numero", "Serie", "Emitente (CNPJ)",
                    "Emitente (nome)", "Data emissao", "Valor (SEFAZ)", "Situacao"])
    for reg in resultado.faltantes_no_sped:
        ws.append([
            reg.chave_normalizada, reg.numero, reg.serie,
            reg.cnpj_emitente_da_chave, reg.emitente_nome, _data(reg.dt_emissao),
            _num(reg.valor), reg.situacao or "Autorizada",
        ])
    _formatar_valores(ws, coluna_valor=7)
    _ajustar_larguras(ws, [46, 10, 8, 20, 34, 13, 15, 14])

    # ---------------- Aba Canceladas escrituradas ----------------
    ws = wb.create_sheet("Canceladas escrituradas")
    _cabecalho(ws, ["Chave de acesso", "Numero", "Emitente", "Situacao na SEFAZ"])
    for c in resultado.canceladas_escrituradas:
        ws.append([c.chave, c.numero, c.emitente, c.situacao_sefaz])
    _ajustar_larguras(ws, [46, 10, 34, 20])

    # ---------------- Aba Divergencias de valor ----------------
    ws = wb.create_sheet("Divergencias de valor")
    _cabecalho(ws, ["Chave de acesso", "Numero", "Emitente",
                    "Valor SEFAZ", "Valor SPED", "Diferenca"])
    for d in resultado.divergencias_valor:
        ws.append([d.chave, d.numero, d.emitente,
                   _num(d.valor_sefaz), _num(d.valor_sped), _num(d.diferenca)])
    _formatar_valores(ws, coluna_valor=4, ate_coluna=6)
    _ajustar_larguras(ws, [46, 10, 34, 15, 15, 15])

    # ---------------- Aba Apenas no SPED ----------------
    ws = wb.create_sheet("Apenas no SPED")
    _cabecalho(ws, ["Chave de acesso", "Numero", "Serie", "Fornecedor",
                    "Data emissao", "Valor (SPED)"])
    for nota in resultado.apenas_no_sped:
        forn = nota.participante.nome if nota.participante else ""
        ws.append([nota.chave_normalizada, nota.numero, nota.serie, forn,
                   _data(nota.dt_emissao), _num(nota.valor_documento)])
    _formatar_valores(ws, coluna_valor=6)
    _ajustar_larguras(ws, [46, 10, 8, 34, 13, 15])

    wb.save(caminho)
    return caminho


def _formatar_valores(ws, coluna_valor: int, ate_coluna: int | None = None) -> None:
    """Aplica formato monetario BR nas colunas de valor (a partir da linha 2)."""
    fim = ate_coluna or coluna_valor
    for row in ws.iter_rows(min_row=2, min_col=coluna_valor, max_col=fim):
        for cel in row:
            cel.number_format = 'R$ #,##0.00'
