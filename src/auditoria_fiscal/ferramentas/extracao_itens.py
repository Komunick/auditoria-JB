"""Item 4 - Extracao de itens das notas para auditoria tributaria.

A partir das notas ja lidas (do SPED ou de XMLs), achata todos os itens em
linhas para conferir a tributacao produto a produto (CFOP, CST, NCM, aliquotas,
ICMS/ST/IPI/PIS/COFINS) e identificar inconsistencias.

Uma unica definicao (CAMPOS) descreve as colunas, usada tanto na exportacao
Excel quanto na previa da interface.
"""

from __future__ import annotations

from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..core.modelos import NotaFiscal


# (chave_interna, titulo, tipo) — tipo controla formatacao: texto|data|num|num4
CAMPOS: list[tuple[str, str, str]] = [
    ("chave", "Chave de acesso", "texto"),
    ("modelo", "Modelo", "texto"),
    ("numero", "Numero", "texto"),
    ("serie", "Serie", "texto"),
    ("data_emissao", "Data emissao", "data"),
    ("operacao", "Operacao", "texto"),
    ("forn_cnpj", "CNPJ emitente", "texto"),
    ("forn_nome", "Emitente / Fornecedor", "texto"),
    ("num_item", "Item", "texto"),
    ("cod_item", "Cod. produto", "texto"),
    ("descricao", "Descricao", "texto"),
    ("ncm", "NCM", "texto"),
    ("cest", "CEST", "texto"),
    ("cfop", "CFOP", "texto"),
    ("cst_icms", "CST/CSOSN", "texto"),
    ("unidade", "Unid.", "texto"),
    ("quantidade", "Quantidade", "num4"),
    ("valor_unitario", "Vlr unitario", "num"),
    ("valor_item", "Vlr total", "num"),
    ("valor_desconto", "Desconto", "num"),
    ("vl_bc_icms", "BC ICMS", "num"),
    ("aliq_icms", "Aliq ICMS %", "num"),
    ("vl_icms", "Vlr ICMS", "num"),
    ("vl_bc_icms_st", "BC ICMS ST", "num"),
    ("aliq_st", "Aliq ST %", "num"),
    ("vl_icms_st", "Vlr ICMS ST", "num"),
    ("cst_ipi", "CST IPI", "texto"),
    ("vl_ipi", "Vlr IPI", "num"),
    ("cst_pis", "CST PIS", "texto"),
    ("vl_pis", "Vlr PIS", "num"),
    ("cst_cofins", "CST COFINS", "texto"),
    ("vl_cofins", "Vlr COFINS", "num"),
]

TITULOS = [titulo for _, titulo, _ in CAMPOS]


def _operacao(nota: NotaFiscal) -> str:
    if nota.ind_oper == "0":
        return "Entrada"
    if nota.ind_oper == "1":
        return "Saida"
    return ""


def extrair_itens(
    notas: list[NotaFiscal],
    somente_operacao: str | None = None,
) -> list[dict]:
    """Achata os itens das notas em linhas (uma por item).

    somente_operacao: None (todas), "0" (entradas) ou "1" (saidas).
    """
    linhas: list[dict] = []
    for nota in notas:
        if somente_operacao is not None and nota.ind_oper != somente_operacao:
            continue
        forn = nota.participante
        base = {
            "chave": nota.chave_normalizada,
            "modelo": nota.modelo,
            "numero": nota.numero,
            "serie": nota.serie,
            "data_emissao": nota.dt_emissao,
            "operacao": _operacao(nota),
            "forn_cnpj": (forn.documento if forn else "") or nota.cnpj_emitente,
            "forn_nome": forn.nome if forn else "",
        }
        for item in nota.itens:
            linha = dict(base)
            linha.update({
                "num_item": item.num_item,
                "cod_item": item.cod_item,
                "descricao": item.descricao,
                "ncm": item.ncm,
                "cest": item.cest,
                "cfop": item.cfop,
                "cst_icms": item.cst_icms,
                "unidade": item.unidade,
                "quantidade": item.quantidade,
                "valor_unitario": item.valor_unitario,
                "valor_item": item.valor_item,
                "valor_desconto": item.valor_desconto,
                "vl_bc_icms": item.vl_bc_icms,
                "aliq_icms": item.aliq_icms,
                "vl_icms": item.vl_icms,
                "vl_bc_icms_st": item.vl_bc_icms_st,
                "aliq_st": item.aliq_st,
                "vl_icms_st": item.vl_icms_st,
                "cst_ipi": item.cst_ipi,
                "vl_ipi": item.vl_ipi,
                "cst_pis": item.cst_pis,
                "vl_pis": item.vl_pis,
                "cst_cofins": item.cst_cofins,
                "vl_cofins": item.vl_cofins,
            })
            linhas.append(linha)
    return linhas


def valor_para_texto(chave: str, tipo: str, valor) -> str:
    """Formata um valor de linha para exibicao (previa da interface)."""
    if valor is None or valor == "":
        return ""
    if tipo == "data":
        return valor.strftime("%d/%m/%Y") if hasattr(valor, "strftime") else str(valor)
    if tipo in ("num", "num4"):
        casas = 4 if tipo == "num4" else 2
        texto = f"{float(valor):,.{casas}f}"
        return texto.replace(",", "X").replace(".", ",").replace("X", ".")
    return str(valor)


def exportar_itens_excel(linhas: list[dict], caminho: str,
                         nome_aba: str = "Itens",
                         filtro_aplicado: str = "") -> str:
    """Exporta as linhas de itens para um arquivo Excel formatado.

    filtro_aplicado: texto opcional exibido acima do cabecalho indicando o
    filtro usado na extracao (ex.: somente documentos de entrada no SPED).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = nome_aba

    linha_cab = 1
    if filtro_aplicado:
        cel = ws.cell(row=1, column=1, value=filtro_aplicado)
        cel.font = Font(bold=True, color="9C874F")  # dourado da logo JB Fraga
        linha_cab = 2

    fonte = Font(bold=True, color="FFFFFF")
    fundo = PatternFill("solid", fgColor="26263A")  # azul-tinta da logo JB Fraga
    for col, titulo in enumerate(TITULOS, start=1):
        cel = ws.cell(row=linha_cab, column=col, value=titulo)
        cel.font = fonte
        cel.fill = fundo
        cel.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = f"A{linha_cab + 1}"

    for linha in linhas:
        valores = []
        for chave, _, tipo in CAMPOS:
            v = linha.get(chave)
            if tipo in ("num", "num4"):
                valores.append(float(v) if isinstance(v, Decimal) else (v or 0))
            elif tipo == "data":
                valores.append(v.strftime("%d/%m/%Y") if hasattr(v, "strftime") else (v or ""))
            else:
                valores.append("" if v is None else str(v))
        ws.append(valores)

    # Formatos numericos por coluna.
    for col, (_, _, tipo) in enumerate(CAMPOS, start=1):
        if tipo in ("num", "num4"):
            fmt = "#,##0.0000" if tipo == "num4" else "#,##0.00"
            letra = get_column_letter(col)
            for cel in ws[letra][linha_cab:]:  # pula o cabecalho (e o filtro)
                cel.number_format = fmt

    # Larguras aproximadas.
    larguras = {"Chave de acesso": 46, "Descricao": 40, "Emitente / Fornecedor": 34,
                "CNPJ emitente": 18, "NCM": 12, "Data emissao": 12}
    for col, titulo in enumerate(TITULOS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = larguras.get(titulo, 13)

    if ws.max_row > linha_cab:
        ws.auto_filter.ref = (f"A{linha_cab}:"
                              f"{get_column_letter(len(CAMPOS))}{ws.max_row}")

    wb.save(caminho)
    return caminho
