"""Exportacao do resultado da comparacao entre dois SPEDs (Item 2)."""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .comparador_sped_sped import ResultadoDiffSped

_AZUL = "1F4E78"
_VERMELHO = "C00000"


def _cabecalho(ws, titulos: list[str]) -> None:
    fonte = Font(bold=True, color="FFFFFF")
    fundo = PatternFill("solid", fgColor=_AZUL)
    for col, titulo in enumerate(titulos, start=1):
        cel = ws.cell(row=1, column=col, value=titulo)
        cel.font = fonte
        cel.fill = fundo
        cel.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def _larguras(ws, larguras: list[int]) -> None:
    for i, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = largura


def gerar_relatorio_diff(resultado: ResultadoDiffSped, caminho: str) -> str:
    wb = Workbook()
    ra, rb = resultado.rotulo_a, resultado.rotulo_b

    # ---------------- Resumo ----------------
    ws = wb.active
    ws.title = "Resumo"
    ws["A1"] = "Comparacao entre versoes de SPED"
    ws["A1"].font = Font(bold=True, size=14, color=_AZUL)
    r = resultado.resumo()
    linhas = [
        (f"Notas em {ra}", r["total_a"]),
        (f"Notas em {rb}", r["total_b"]),
        ("Notas presentes nos dois", r["conciliadas"]),
        ("Notas identicas", r["iguais"]),
        ("Notas com divergencia", r["divergentes"]),
        ("Total de campos divergentes", r["total_diferencas"]),
        (f"Notas apenas em {ra}", r["apenas_em_a"]),
        (f"Notas apenas em {rb}", r["apenas_em_b"]),
    ]
    for i, (rotulo, valor) in enumerate(linhas):
        linha = 3 + i
        ws.cell(row=linha, column=1, value=rotulo).font = Font(bold=True)
        cel = ws.cell(row=linha, column=2, value=valor)
        if rotulo == "Notas com divergencia" and valor:
            cel.font = Font(bold=True, color=_VERMELHO)
    _larguras(ws, [40, 14])

    # ---------------- Divergencias (uma linha por campo) ----------------
    ws = wb.create_sheet("Divergencias")
    _cabecalho(ws, ["Chave de acesso", "Numero", "Fornecedor", "Nivel", "Item",
                    "Campo", ra, rb])
    for nota in resultado.divergentes:
        for d in nota.diferencas:
            ws.append([nota.chave, nota.numero, nota.fornecedor,
                       "Nota" if d.nivel == "nota" else "Item",
                       d.num_item, d.campo, d.valor_a, d.valor_b])
    _larguras(ws, [46, 10, 32, 8, 8, 22, 20, 20])

    # ---------------- Apenas em A / B ----------------
    for titulo, notas in (("Apenas em " + ra, resultado.apenas_em_a),
                          ("Apenas em " + rb, resultado.apenas_em_b)):
        # nome de aba do Excel: max 31 chars
        ws = wb.create_sheet(titulo[:31])
        _cabecalho(ws, ["Chave de acesso", "Numero", "Serie", "Fornecedor", "Valor"])
        for nota in notas:
            forn = nota.participante.nome if nota.participante else ""
            cel_valor = float(nota.valor_documento or 0)
            ws.append([nota.chave_normalizada, nota.numero, nota.serie, forn, cel_valor])
        for cel in ws["E"][1:]:
            cel.number_format = "#,##0.00"
        _larguras(ws, [46, 10, 8, 32, 15])

    wb.save(caminho)
    return caminho
